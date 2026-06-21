import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app
from app import db
from app.models import (
    Usuario, Calificacion, AlertaDesvio, Justificacion,
    RegistroParada, Turno, Ruta, ReporteGenerado
)


# Paleta de colores del sistema
COLOR_HEADER = "1A3A5C"   # Azul oscuro
COLOR_OK     = "27AE60"   # Verde
COLOR_ALERTA = "E74C3C"   # Rojo
COLOR_WARN   = "F39C12"   # Naranja
COLOR_LIGHT  = "EBF5FB"   # Azul claro para filas alternas


def _estilo_header(ws, fila, columnas, texto_columnas):
    """Aplica estilo de encabezado a una fila."""
    for col, texto in zip(range(1, columnas + 1), texto_columnas):
        cell = ws.cell(row=fila, column=col, value=texto)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin", color="FFFFFF")
        )


def _autoajustar(ws):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generar_reporte_cumplimiento(desde=None, hasta=None, ruta_id=None, generado_por_id=None):
    """
    Reporte de cumplimiento de rutas: paradas OK vs totales por día.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cumplimiento de Rutas"

    # Título
    ws.merge_cells("A1:G1")
    titulo = ws["A1"]
    titulo.value = f"REPORTE DE CUMPLIMIENTO DE RUTAS — SIG Transporte"
    titulo.font = Font(bold=True, size=14, color=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    subtitulo = ws["A2"]
    subtitulo.value = f"Período: {desde or 'Inicio'} al {hasta or 'Hoy'}"
    subtitulo.alignment = Alignment(horizontal="center")

    # Encabezados
    headers = ["Fecha", "Placa", "Ruta", "Chofer", "Paradas OK", "Total Paradas", "% Cumplimiento"]
    _estilo_header(ws, 4, len(headers), headers)

    # Consulta
    query = db.session.query(
        RegistroParada, Turno, Ruta, Usuario
    ).join(Turno, RegistroParada.turno_id == Turno.id)\
     .join(Ruta, Turno.ruta_id == Ruta.id)\
     .join(Usuario, RegistroParada.chofer_id == Usuario.id)

    if desde:
        query = query.filter(RegistroParada.registrado_en >= desde)
    if hasta:
        query = query.filter(RegistroParada.registrado_en <= hasta)
    if ruta_id:
        query = query.filter(Turno.ruta_id == ruta_id)

    registros = query.all()

    # Agrupar por turno
    turnos_data = {}
    for reg, turno, ruta, chofer in registros:
        key = turno.id
        if key not in turnos_data:
            turnos_data[key] = {
                "fecha": turno.fecha,
                "placa": turno.minibus.placa if turno.minibus else "—",
                "ruta": ruta.nombre,
                "chofer": f"{chofer.nombre} {chofer.apellido}",
                "ok": 0,
                "total": 0,
            }
        turnos_data[key]["total"] += 1
        if reg.en_ruta:
            turnos_data[key]["ok"] += 1

    fila = 5
    for i, (turno_id, datos) in enumerate(sorted(turnos_data.items())):
        pct = round((datos["ok"] / datos["total"]) * 100, 1) if datos["total"] > 0 else 0
        color_fila = COLOR_LIGHT if i % 2 == 0 else "FFFFFF"

        valores = [
            str(datos["fecha"]),
            datos["placa"],
            datos["ruta"],
            datos["chofer"],
            datos["ok"],
            datos["total"],
            f"{pct}%",
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color_fila)
            cell.alignment = Alignment(horizontal="center")
            if col == 7:
                cell.font = Font(
                    bold=True,
                    color=COLOR_OK if pct >= 80 else (COLOR_WARN if pct >= 60 else COLOR_ALERTA)
                )
        fila += 1

    _autoajustar(ws)
    return _guardar_excel(wb, "cumplimiento", generado_por_id)


def generar_reporte_calificaciones(periodo=None, generado_por_id=None):
    """
    Reporte de calificaciones por chofer en un período.
    """
    periodo = periodo or datetime.utcnow().strftime("%Y-%m")
    wb = Workbook()
    ws = wb.active
    ws.title = "Calificaciones Choferes"

    ws.merge_cells("A1:H1")
    titulo = ws["A1"]
    titulo.value = f"CALIFICACIONES DE CHOFERES — Período {periodo}"
    titulo.font = Font(bold=True, size=14, color=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center")

    headers = ["Chofer", "CI", "Período", "Paradas OK", "Total Paradas",
               "Alertas", "Faltas Justif.", "Calificación", "Nivel"]
    _estilo_header(ws, 3, len(headers), headers)

    calificaciones = Calificacion.query.filter_by(periodo=periodo)\
        .join(Usuario, Calificacion.chofer_id == Usuario.id).all()

    fila = 4
    for i, cal in enumerate(calificaciones):
        color_fila = COLOR_LIGHT if i % 2 == 0 else "FFFFFF"
        valores = [
            f"{cal.chofer.nombre} {cal.chofer.apellido}" if cal.chofer else "—",
            cal.chofer.numero_ci if cal.chofer else "—",
            cal.periodo,
            cal.paradas_ok,
            cal.total_paradas,
            cal.alertas_generadas,
            cal.faltas_justificadas,
            float(cal.calificacion) if cal.calificacion else 0,
            cal.nivel or "—",
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color_fila)
            cell.alignment = Alignment(horizontal="center")
        fila += 1

    _autoajustar(ws)
    return _guardar_excel(wb, "calificaciones", generado_por_id)


def generar_reporte_alertas(desde=None, hasta=None, chofer_id=None, generado_por_id=None):
    """
    Reporte de alertas de desvío y su resolución.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas de Desvío"

    ws.merge_cells("A1:I1")
    titulo = ws["A1"]
    titulo.value = "REPORTE DE ALERTAS DE DESVÍO — SIG Transporte"
    titulo.font = Font(bold=True, size=14, color=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center")

    headers = ["Fecha/Hora", "Placa", "Chofer", "Distancia (m)", "Disparada por",
               "Estado", "Motivo Justif.", "Decisión Admin", "¿Afecta calificación?"]
    _estilo_header(ws, 3, len(headers), headers)

    query = AlertaDesvio.query
    if desde:
        query = query.filter(AlertaDesvio.generada_en >= desde)
    if hasta:
        query = query.filter(AlertaDesvio.generada_en <= hasta)
    if chofer_id:
        query = query.filter(AlertaDesvio.chofer_id == chofer_id)

    alertas = query.order_by(AlertaDesvio.generada_en.desc()).all()

    fila = 4
    for i, alerta in enumerate(alertas):
        j = alerta.justificacion
        afecta = "Sí" if alerta.estado in ["sin_respuesta", "rechazada"] else "No"
        color_fila = COLOR_LIGHT if i % 2 == 0 else "FFFFFF"

        valores = [
            alerta.generada_en.strftime("%Y-%m-%d %H:%M") if alerta.generada_en else "—",
            alerta.minibus.placa if alerta.minibus else "—",
            f"{alerta.chofer.nombre} {alerta.chofer.apellido}" if alerta.chofer else "—",
            float(alerta.distancia_metros) if alerta.distancia_metros else 0,
            alerta.disparada_por,
            alerta.estado,
            j.motivo if j else "Sin justificación",
            j.estado if j else "—",
            afecta,
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color_fila)
            cell.alignment = Alignment(horizontal="center")
            if col == 9:
                cell.font = Font(
                    bold=True,
                    color=COLOR_ALERTA if afecta == "Sí" else COLOR_OK
                )
        fila += 1

    _autoajustar(ws)
    return _guardar_excel(wb, "alertas", generado_por_id)


def _guardar_excel(wb, tipo, generado_por_id=None):
    """Guarda el archivo Excel y registra en BD."""
    carpeta = current_app.config.get("REPORTS_FOLDER", "./static/reports")
    os.makedirs(carpeta, exist_ok=True)
    nombre = f"{tipo}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_archivo = os.path.join(carpeta, nombre)
    wb.save(ruta_archivo)

    # Registrar en BD
    if generado_por_id:
        from app import db as _db
        reporte = ReporteGenerado(
            generado_por=generado_por_id,
            tipo=tipo,
            formato="excel",
            archivo_url=f"/static/reports/{nombre}",
        )
        _db.session.add(reporte)
        _db.session.commit()

    return ruta_archivo, f"/static/reports/{nombre}"
